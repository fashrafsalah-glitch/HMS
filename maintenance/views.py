from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseBadRequest, JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from django.utils import timezone
from django.db.models import Q
from django.template.loader import render_to_string
import json
from .forms import CompanyForm, DeviceFormBasic, DeviceTransferForm, DeviceTypeForm, DeviceUsageForm, DeviceAccessoryForm, DeviceCategoryForm, DeviceSubCategoryForm
from .models import Company, Device, DeviceTransferRequest, DeviceType, DeviceUsage, DeviceCleaningLog, DeviceSterilizationLog, DeviceMaintenanceLog, DeviceCategory, DeviceSubCategory, PreventiveMaintenanceSchedule, WorkOrder, JobPlan
from manager.models import Department, Room, Bed, Patient

from .models import DeviceAccessory, AccessoryTransaction
from .forms import AccessoryScanForm

@login_required
def clean_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    device.clean_status = 'clean'
    device.last_cleaned_by = request.user
    device.last_cleaned_at = timezone.now()
    device.save()
        # إضافة سجل جديد
    DeviceCleaningLog.objects.create(device=device, cleaned_by=request.user)


    messages.success(request, f"✅ تم تنظيف الجهاز بواسطة {request.user.username}")
    return redirect('maintenance:device_detail', pk=device.id)

@login_required
def cleaning_history(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    logs = device.cleaning_logs.all().order_by('-cleaned_at')
    return render(request, 'maintenance/cleaning_history.html', {'device': device, 'logs': logs})

@login_required
def sterilize_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    device.sterilization_status = 'sterilized'
    device.last_sterilized_by = request.user
    device.last_sterilized_at = timezone.now()
    device.save()

    # إضافة سجل جديد
    DeviceSterilizationLog.objects.create(device=device, sterilized_by=request.user)
    messages.success(request, f"✅ تم تعقيم الجهاز بواسطة {request.user.username}")
    return redirect('maintenance:device_detail', pk=device.id)

@login_required
def sterilization_history(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    logs = device.sterilization_logs.all().order_by('-sterilized_at')
    return render(request, 'maintenance/sterilization_history.html', {'device': device, 'logs': logs})

@login_required
def perform_maintenance(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    device.status = 'working'
    device.last_maintained_by = request.user
    device.last_maintained_at = timezone.now()
    device.save()

    # إضافة سجل جديد
    DeviceMaintenanceLog.objects.create(device=device, maintained_by=request.user)
    messages.success(request, f"✅ تم صيانة الجهاز بواسطة {request.user.username}")
    return redirect('maintenance:device_detail', pk=device.id)

@login_required
def maintenance_history(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    logs =  device.maintenance_logs.all().order_by('-maintained_at')
    return render(request, 'maintenance/maintenance_history.html', {'device': device, 'logs': logs})

# AJAX Views for instant device status updates
@login_required
@require_POST
def ajax_clean_device(request, device_id):
    """AJAX endpoint for instant device cleaning"""
    try:
        device = get_object_or_404(Device, id=device_id)
        device.clean_status = 'clean'
        device.last_cleaned_by = request.user
        device.last_cleaned_at = timezone.now()
        device.save()
        
        # Add cleaning log
        DeviceCleaningLog.objects.create(device=device, cleaned_by=request.user)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ تم تنظيف الجهاز بواسطة {request.user.username}',
            'clean_status': device.get_clean_status_display(),
            'clean_status_class': 'success',
            'last_cleaned_at': device.last_cleaned_at.strftime('%Y-%m-%d %H:%M'),
            'last_cleaned_by': str(device.last_cleaned_by)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في تنظيف الجهاز: {str(e)}'
        })

@login_required
@require_POST
def ajax_sterilize_device(request, device_id):
    """AJAX endpoint for instant device sterilization"""
    try:
        device = get_object_or_404(Device, id=device_id)
        device.sterilization_status = 'sterilized'
        device.last_sterilized_by = request.user
        device.last_sterilized_at = timezone.now()
        device.save()
        
        # Add sterilization log
        DeviceSterilizationLog.objects.create(device=device, sterilized_by=request.user)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ تم تعقيم الجهاز بواسطة {request.user.username}',
            'sterilization_status': device.get_sterilization_status_display(),
            'sterilization_status_class': 'success',
            'last_sterilized_at': device.last_sterilized_at.strftime('%Y-%m-%d %H:%M'),
            'last_sterilized_by': str(device.last_sterilized_by)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في تعقيم الجهاز: {str(e)}'
        })

@login_required
@require_POST
def ajax_perform_maintenance(request, device_id):
    """AJAX endpoint for instant device maintenance"""
    try:
        device = get_object_or_404(Device, id=device_id)
        device.status = 'working'
        device.last_maintained_by = request.user
        device.last_maintained_at = timezone.now()
        device.save()
        
        # Add maintenance log
        DeviceMaintenanceLog.objects.create(device=device, maintained_by=request.user)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ تم فحص الجهاز بواسطة {request.user.username}',
            'status': device.get_status_display(),
            'status_class': 'success',
            'last_maintained_at': device.last_maintained_at.strftime('%Y-%m-%d %H:%M'),
            'last_maintained_by': str(device.last_maintained_by)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في فحص الجهاز: {str(e)}'
        })

def assign_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    
    # Get all patients - simplified approach
    patients = Patient.objects.all().order_by('first_name', 'last_name')

    if request.method == 'POST':
        patient_id = request.POST.get('patient')

        if not patient_id:
            messages.error(request, "لم يتم اختيار مريض.")
            return render(request, 'maintenance/assign_device.html', {'device': device, 'patients': patients})

        try:
            patient = Patient.objects.get(id=patient_id)
        except Patient.DoesNotExist:
            messages.error(request, "المريض المحدد غير موجود.")
            return render(request, 'maintenance/assign_device.html', {'device': device, 'patients': patients})

        device.current_patient = patient
        device.status = 'in_use'
        device.save()
        messages.success(request, f'تم ربط الجهاز "{device.name}" بالمريض "{patient.first_name} {patient.last_name}" بنجاح.')
        return redirect('maintenance:device_list')

    return render(request, 'maintenance/assign_device.html', {'device': device, 'patients': patients})

def release_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    device.current_patient = None
    device.status = 'needs_check'
    device.clean_status = 'needs_cleaning'
    device.sterilization_status = 'needs_sterilization'
    device.save()
    return redirect('maintenance:device_list')



@login_required
def device_transfer_list(request):
    """
    عرض قائمة طلبات نقل الأجهزة للموافقة عليها
    """
    transfer_requests = DeviceTransferRequest.objects.all().order_by('-requested_at')
    
    context = {
        'transfer_requests': transfer_requests,
    }
    return render(request, 'maintenance/device_transfer_list.html', context)


@login_required
def approve_transfer(request, transfer_id):
    transfer = get_object_or_404(DeviceTransferRequest, id=transfer_id)

    if request.method == 'POST' and not transfer.is_approved:
        try:
            # تنفيذ النقل
            device = transfer.device
            old_dept = device.department
            old_room = device.room
            
            device.department = transfer.to_department
            device.room = transfer.to_room
            device.save()

            transfer.is_approved = True
            transfer.approved_by = request.user
            transfer.approved_at = timezone.now()
            transfer.save()
            
            messages.success(request, f'تم الموافقة على نقل الجهاز {device.name} من {old_dept} غرفة {old_room} إلى {transfer.to_department.name} غرفة {transfer.to_room.number}')
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء الموافقة على النقل: {str(e)}')
    elif transfer.is_approved:
        messages.info(request, 'تم الموافقة على هذا الطلب مسبقاً')

    return redirect('maintenance:device_transfer_list')

@require_GET
def load_rooms(request):
    department_id = request.GET.get('department_id')
    rooms = Room.objects.filter(department_id=department_id).values('id', 'number')
    return JsonResponse(list(rooms), safe=False)

def device_list(request):
    devices = Device.objects.all()

    # فلترة حسب التصنيف
    category_id = request.GET.get('category')
    if category_id:
        devices = devices.filter(category_id=category_id)

    # فلترة حسب التصنيف الفرعي
    subcategory_id = request.GET.get('subcategory')
    if subcategory_id:
        devices = devices.filter(subcategory_id=subcategory_id)

    # فلترة حسب القسم
    department_id = request.GET.get('department')
    if department_id:
        devices = devices.filter(department_id=department_id)

    # فلترة حسب الغرفة
    room_id = request.GET.get('room')
    if room_id:
        devices = devices.filter(room_id=room_id)

    categories = DeviceCategory.objects.all()
    subcategories = DeviceSubCategory.objects.all()
    departments = Department.objects.all()
    rooms = Room.objects.all()

    return render(request, 'maintenance/device_list.html', {
        'devices': devices,
        'categories': categories,
        'subcategories': subcategories,
        'departments': departments,
        'rooms': rooms,
    })

def load_subcategories(request):
    category_id = request.GET.get('category_id')
    subcategories = DeviceSubCategory.objects.filter(category_id=category_id).values('id', 'name')
    return JsonResponse(list(subcategories), safe=False)

# تم دمج هذه الوظيفة مع الوظيفة السابقة
# def device_list(request):
#    category_id = request.GET.get('category')
#    subcategory_id = request.GET.get('subcategory')
#    department_id = request.GET.get('department')
#    room_id = request.GET.get('room')
#
#    devices = Device.objects.all()

# تم تعليق الكود المتبقي من الوظيفة المكررة
#    if category_id:
#        devices = devices.filter(category_id=category_id)
#    if subcategory_id:
#        devices = devices.filter(subcategory_id=subcategory_id)
#    if department_id:
#        devices = devices.filter(department_id=department_id)
#    if room_id:
#        devices = devices.filter(room_id=room_id)

# تم تعليق الكود المتبقي من الوظيفة المكررة
#    context = {
#        'devices': devices,
#        'categories': DeviceCategory.objects.all(),
#        'subcategories': DeviceSubCategory.objects.all(),
#        'departments': Department.objects.all(),
#        'rooms': Room.objects.all(),
#    }
#    return render(request, 'maintenance/device_list.html', context)

def add_device_category(request):
    if request.method == 'POST':
        form = DeviceCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('maintenance:add_device')  # الرجوع إلى إضافة جهاز
    else:
        form = DeviceCategoryForm()
    return render(request, 'maintenance/device_category_form.html', {'form': form})


def add_device_subcategory(request):
    if request.method == 'POST':
        form = DeviceSubCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('maintenance:add_device')  # الرجوع إلى إضافة جهاز
    else:
        form = DeviceSubCategoryForm()
    return render(request, 'maintenance/device_subcategory_form.html', {'form': form})
def device_detail(request, pk):
    device = get_object_or_404(Device, pk=pk)
    # Keep device detail page focused on device info only
    context = {
        'device': device,
    }
    return render(request, 'maintenance/device_detail.html', context)

@login_required
def redirect_to_accessories(request, pk):
    """Redirect old add-accessory URL to device detail page with accessories tab"""
    return redirect('maintenance:device_accessories', device_id=pk)

def device_edit(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        form = DeviceFormBasic(request.POST, instance=device)
        if form.is_valid():
            form.save()
            return redirect('maintenance:device_list')
    else:
        form = DeviceFormBasic(instance=device)
    return render(request, 'maintenance/device_form.html', {'form': form, 'device': device})

def device_delete(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        device.delete()
        return redirect('maintenance:device_list')
    return render(request, 'maintenance/device_confirm_delete.html', {'device': device})

def transfer_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    
    # Check device transfer eligibility
    transfer_errors = []
    if hasattr(device, 'status') and device.status != 'working':
        transfer_errors.append(f"الجهاز في حالة: {device.get_status_display()} - يجب أن يكون يعمل")
    if hasattr(device, 'clean_status') and device.clean_status != 'clean':
        transfer_errors.append("الجهاز يحتاج تنظيف قبل النقل")
    if hasattr(device, 'sterilization_status') and device.sterilization_status != 'sterilized':
        transfer_errors.append("الجهاز يحتاج تعقيم قبل النقل")
    if hasattr(device, 'availability') and not device.availability:
        transfer_errors.append("الجهاز غير متاح حالياً - لا يمكن نقله")

    if request.method == 'POST':
        form = DeviceTransferForm(request.POST)
        if form.is_valid():
            if transfer_errors:
                for error in transfer_errors:
                    messages.error(request, error)
                return render(request, 'maintenance/device_transfer.html', {
                    'form': form, 
                    'device': device, 
                    'transfer_errors': transfer_errors
                })
            
            transfer = DeviceTransferRequest.objects.create(
                device=device,
                from_department=device.department,
                to_department=form.cleaned_data['to_department'],
                from_room=device.room,
                to_room=form.cleaned_data['to_room'],
                requested_by=request.user,
            )
            messages.success(request, "تم إرسال طلب نقل الجهاز بنجاح")
            return redirect('maintenance:device_detail', pk=device.id)
    else:
        form = DeviceTransferForm(initial={
            'to_department': device.department,
            'to_room': device.room
        })

    return render(request, 'maintenance/device_transfer.html', {
        'form': form, 
        'device': device, 
        'transfer_errors': transfer_errors
    })



def approve_transfer(request, transfer_id):
    transfer = get_object_or_404(DeviceTransferRequest, id=transfer_id)
    device = transfer.device
    
    # Re-check device eligibility before approval
    transfer_errors = []
    if hasattr(device, 'status') and device.status != 'working':
        transfer_errors.append(f"الجهاز في حالة: {device.get_status_display()} - يجب أن يكون يعمل")
    if hasattr(device, 'clean_status') and device.clean_status != 'clean':
        transfer_errors.append("الجهاز يحتاج تنظيف قبل النقل")
    if hasattr(device, 'sterilization_status') and device.sterilization_status != 'sterilized':
        transfer_errors.append("الجهاز يحتاج تعقيم قبل النقل")
    if hasattr(device, 'availability') and not device.availability:
        transfer_errors.append("الجهاز غير متاح حالياً - لا يمكن نقله")

    if request.method == 'POST':
        if transfer_errors:
            for error in transfer_errors:
                messages.error(request, f"لا يمكن الموافقة على النقل: {error}")
            return redirect('maintenance:device_detail', pk=device.id)
        
        # Approve transfer
        transfer.is_approved = True
        transfer.approved_by = request.user
        transfer.approved_at = timezone.now()
        transfer.save()

        # Create transfer log
        DeviceTransferLog.objects.create(
            device=device,
            from_department=transfer.from_department,
            from_room=transfer.from_room,
            to_department=transfer.to_department,
            to_room=transfer.to_room,
            moved_by=request.user,
            note=f"نقل معتمد - طلب رقم {transfer.id}"
        )

        # Update device location
        device.department = transfer.to_department
        device.room = transfer.to_room
        device.save()

        messages.success(request, "تم قبول نقل الجهاز بنجاح.")
        return redirect('maintenance:device_detail', pk=device.id)
    
    return redirect('maintenance:device_detail', pk=device.id)

def index(request):
    # إحصائيات صيانة الأجهزة
    total_devices = Device.objects.count()
    working_devices = Device.objects.filter(status='working').count()
    needs_maintenance = Device.objects.filter(status='needs_maintenance').count()
    out_of_order = Device.objects.filter(status='out_of_order').count()
    needs_cleaning = Device.objects.filter(clean_status='needs_cleaning').count()
    needs_sterilization = Device.objects.filter(sterilization_status='needs_sterilization').count()
    
    # إحصائيات الصيانة العامة (CMMS)
    try:
        from .models import WorkOrder, ServiceRequest
        total_work_orders = WorkOrder.objects.count()
        pending_work_orders = WorkOrder.objects.filter(status='pending').count()
        total_service_requests = ServiceRequest.objects.count()
        pending_service_requests = ServiceRequest.objects.filter(status='pending').count()
    except ImportError:
        total_work_orders = 0
        pending_work_orders = 0
        total_service_requests = 0
        pending_service_requests = 0
    
    # الأجهزة الأكثر استخداماً
    popular_devices = Device.objects.filter(current_patient__isnull=False)[:5]
    
    # آخر عمليات الصيانة
    recent_maintenance = DeviceMaintenanceLog.objects.select_related('device', 'maintained_by').order_by('-maintained_at')[:5]
    
    context = {
        # إحصائيات الأجهزة
        'total_devices': total_devices,
        'working_devices': working_devices,
        'needs_maintenance': needs_maintenance,
        'out_of_order': out_of_order,
        'needs_cleaning': needs_cleaning,
        'needs_sterilization': needs_sterilization,
        
        # إحصائيات CMMS
        'total_work_orders': total_work_orders,
        'pending_work_orders': pending_work_orders,
        'total_service_requests': total_service_requests,
        'pending_service_requests': pending_service_requests,
        
        # بيانات إضافية
        'popular_devices': popular_devices,
        'recent_maintenance': recent_maintenance,
    }
    
    return render(request, 'maintenance/index.html', context)



def add_device(request):
    if request.method == 'POST':
        form = DeviceFormBasic(request.POST)
        if form.is_valid():
            form.save()
            return redirect('maintenance:device_list')
        else:
            print("Form Errors:", form.errors)  # ← اطبع الأخطاء هنا
    else:
        form = DeviceFormBasic()
    return render(request, 'maintenance/add_device.html', {'form': form})





def device_info(request, pk):
    device = get_object_or_404(Device, pk=pk)
    return render(request, 'maintenance/device_info.html', {'device': device})

def add_accessory(request, pk):
    device = get_object_or_404(Device, pk=pk)
    
    if request.method == 'POST':
        form = DeviceAccessoryForm(request.POST)
        if form.is_valid():
            accessory = form.save(commit=False)
            accessory.device = device
            accessory.save()
            messages.success(request, f'تم إضافة الملحق "{accessory.name}" بنجاح للجهاز {device.name}')
            return redirect('maintenance:device_info', pk=device.id)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = DeviceAccessoryForm()
    
    return render(request, 'maintenance/add_accessory.html', {
        'device': device,
        'form': form
    })

def maintenance_schedule(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    
    # جلب الصيانات المجدولة للجهاز
    scheduled_maintenances = PreventiveMaintenanceSchedule.objects.filter(device=device)
    
    # جلب سجل الصيانة من DeviceMaintenanceLog
    maintenance_logs = DeviceMaintenanceLog.objects.filter(device=device).order_by('-maintained_at')[:10]
    
    # جلب أوامر العمل المرتبطة بالجهاز
    work_orders = WorkOrder.objects.filter(service_request__device=device).order_by('-created_at')[:5]
    
    if request.method == 'POST':
        # معالجة بيانات النموذج
        schedule_type = request.POST.get('schedule_type')
        next_maintenance = request.POST.get('next_maintenance')
        maintenance_type = request.POST.get('maintenance_type')
        maintenance_notes = request.POST.get('maintenance_notes')
        
        # إنشاء جدولة صيانة جديدة
        from datetime import datetime
        try:
            next_date = datetime.strptime(next_maintenance, '%Y-%m-%d').date()
            
            # إنشاء أو الحصول على خطة عمل افتراضية
            job_plan, created = JobPlan.objects.get_or_create(
                name=f"خطة صيانة {maintenance_type}",
                defaults={
                    'description': f"خطة صيانة {maintenance_type} افتراضية",
                    'estimated_duration': 60,  # 60 دقيقة افتراضي
                    'created_by': request.user
                }
            )
            
            PreventiveMaintenanceSchedule.objects.create(
                device=device,
                job_plan=job_plan,
                frequency=schedule_type,
                next_due_date=next_date,
                created_by=request.user
            )
            messages.success(request, f'تم جدولة صيانة {maintenance_type} للجهاز {device.name} بتاريخ {next_maintenance}')
        except ValueError:
            messages.error(request, 'تاريخ غير صحيح')
        
        return redirect('maintenance:maintenance_schedule', device_id=device_id)
    
    return render(request, 'maintenance/maintenance_schedule.html', {
        'device': device,
        'device_id': device_id,
        'scheduled_maintenances': scheduled_maintenances,
        'maintenance_logs': maintenance_logs,
        'work_orders': work_orders
    })

@login_required
def edit_schedule(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    
    if request.method == 'POST':
        schedule_id = request.POST.get('schedule_id')
        schedule_type = request.POST.get('schedule_type')
        next_maintenance = request.POST.get('next_maintenance')
        
        try:
            schedule = get_object_or_404(PreventiveMaintenanceSchedule, id=schedule_id, device=device)
            
            from datetime import datetime
            next_date = datetime.strptime(next_maintenance, '%Y-%m-%d').date()
            
            schedule.frequency = schedule_type
            schedule.next_due_date = next_date
            schedule.save()
            
            messages.success(request, 'تم تحديث الصيانة المجدولة بنجاح')
        except ValueError:
            messages.error(request, 'تاريخ غير صحيح')
        except Exception as e:
            messages.error(request, 'حدث خطأ أثناء التحديث')
        return redirect('maintenance:maintenance_schedule', device_id=device_id)

@login_required
def delete_schedule(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    
    if request.method == 'POST':
        schedule_id = request.POST.get('schedule_id')
        
        try:
            schedule = get_object_or_404(PreventiveMaintenanceSchedule, id=schedule_id, device=device)
            schedule_name = schedule.job_plan.name
            schedule.delete()
            
            messages.success(request, f'تم حذف الصيانة المجدولة "{schedule_name}" بنجاح')
        except Exception as e:
            messages.error(request, 'حدث خطأ أثناء الحذف')
    
    return redirect('maintenance:maintenance_schedule', device_id=device_id)

# Device Operations Views
def clean_device(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        DeviceCleaningLog.objects.create(
            device=device,
            cleaned_by=request.user,
            notes=notes
        )
        messages.success(request, f'تم تسجيل تنظيف الجهاز "{device.name}" بنجاح')
        return redirect('maintenance:device_info', pk=device.id)
    
    return render(request, 'maintenance/clean_device.html', {'device': device})

def sterilize_device(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        method = request.POST.get('method', '')
        DeviceSterilizationLog.objects.create(
            device=device,
            sterilized_by=request.user,
            notes=notes,
            method=method
        )
        messages.success(request, f'تم تسجيل تعقيم الجهاز "{device.name}" بنجاح')
        return redirect('maintenance:device_info', pk=device.id)
    
    return render(request, 'maintenance/sterilize_device.html', {'device': device})

def perform_maintenance(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        maintenance_type = request.POST.get('maintenance_type', '')
        DeviceMaintenanceLog.objects.create(
            device=device,
            maintained_by=request.user,
            description=notes,
            maintenance_type=maintenance_type
        )
        messages.success(request, f'تم تسجيل صيانة الجهاز "{device.name}" بنجاح')
        return redirect('maintenance:device_info', pk=device.id)
    
    return render(request, 'maintenance/perform_maintenance.html', {'device': device})

# History Views
def cleaning_history(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    cleaning_logs = DeviceCleaningLog.objects.filter(device=device).order_by('-cleaned_at')
    return render(request, 'maintenance/cleaning_history.html', {
        'device': device,
        'cleaning_logs': cleaning_logs
    })

def sterilization_history(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    sterilization_logs = DeviceSterilizationLog.objects.filter(device=device).order_by('-sterilized_at')
    return render(request, 'maintenance/sterilization_history.html', {
        'device': device,
        'sterilization_logs': sterilization_logs
    })

def maintenance_history(request, device_id):
    device = get_object_or_404(Device, pk=device_id)
    maintenance_logs = DeviceMaintenanceLog.objects.filter(device=device).order_by('-maintained_at')
    return render(request, 'maintenance/maintenance_history.html', {
        'device': device,
        'maintenance_logs': maintenance_logs
    })

def add_emergency_request(request, pk):
    return render(request, 'maintenance/add_emergency_request.html', {'device_id': pk})

def add_spare_part(request, pk):
    return render(request, 'maintenance/add_spare_part.html', {'device_id': pk})





def department_devices(request, department_id):
    department = get_object_or_404(Department, id=department_id)

    # استبعاد الأجهزة التي تم طلب نقلها لهذا القسم ولم تُقبل بعد
    pending_transfers = DeviceTransferRequest.objects.filter(
        to_department=department,
        is_approved=False
    ).select_related('device')

    pending_devices_ids = [t.device.id for t in pending_transfers]

    # عرض الأجهزة الفعلية فقط التي لا تنتظر موافقة النقل
    actual_devices = Device.objects.filter(
        department=department
    ).exclude(id__in=pending_devices_ids)

    return render(request, 'maintenance/department_device_list.html', {
        'department': department,
        'actual_devices': actual_devices,
        'pending_transfers': pending_transfers
    })

from .models import DeviceTransferRequest

def device_transfer_history(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    transfers = DeviceTransferRequest.objects.filter(device=device).select_related(
        'from_department', 'to_department', 'from_room', 'to_room', 'requested_by', 'approved_by'
    ).order_by('-requested_at')

    return render(request, 'maintenance/device_transfer_history.html', {
        'device': device,
        'transfers': transfers
    })


def get_rooms(request):
    department_id = request.GET.get('department_id')
    rooms = Room.objects.filter(department_id=department_id).order_by('number')
    html = render_to_string('maintenance/partials/room_dropdown.html', {'rooms': rooms})
    return JsonResponse({'html': html})

def get_beds(request):
    room_id = request.GET.get('room_id')
    beds = Bed.objects.filter(room_id=room_id).order_by('bed_number')
    html = render_to_string('maintenance/partials/bed_dropdown.html', {'beds': beds})
    return JsonResponse({'html': html})


def add_company(request):
    form = CompanyForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('maintenance:device_list')
    return render(request, 'maintenance/add_company.html', {'form': form})

def add_device_type(request):
    form = DeviceTypeForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('maintenance:device_list')
    return render(request, 'maintenance/add_device_type.html', {'form': form})

def add_device_usage(request):
    form = DeviceUsageForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('maintenance:device_list')
    return render(request, 'maintenance/add_device_usage.html', {'form': form})


# تعديل شركة
def edit_company(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            return redirect('company_list')
    else:
        form = CompanyForm(instance=company)
    return render(request, 'maintenance/edit_company.html', {'form': form})

# حذف شركة
def delete_company(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        company.delete()
        return redirect('company_list')
    return render(request, 'maintenance/confirm_delete.html', {'object': company, 'type': 'Company'})

# تعديل نوع الجهاز
def edit_device_type(request, pk):
    dtype = get_object_or_404(DeviceType, pk=pk)
    if request.method == 'POST':
        form = DeviceTypeForm(request.POST, instance=dtype)
        if form.is_valid():
            form.save()
            return redirect('device_type_list')
    else:
        form = DeviceTypeForm(instance=dtype)
    return render(request, 'maintenance/edit_device_type.html', {'form': form})

# حذف نوع الجهاز
def delete_device_type(request, pk):
    dtype = get_object_or_404(DeviceType, pk=pk)
    if request.method == 'POST':
        dtype.delete()
        return redirect('device_type_list')
    return render(request, 'maintenance/confirm_delete.html', {'object': dtype, 'type': 'Device Type'})

# تعديل استخدام الجهاز
def edit_device_usage(request, pk):
    usage = get_object_or_404(DeviceUsage, pk=pk)
    if request.method == 'POST':
        form = DeviceUsageForm(request.POST, instance=usage)
        if form.is_valid():
            form.save()
            return redirect('device_usage_list')
    else:
        form = DeviceUsageForm(instance=usage)
    return render(request, 'maintenance/edit_device_usage.html', {'form': form})

# حذف استخدام الجهاز
def delete_device_usage(request, pk):
    usage = get_object_or_404(DeviceUsage, pk=pk)
    if request.method == 'POST':
        usage.delete()
        return redirect('device_usage_list')
    return render(request, 'maintenance/confirm_delete.html', {'object': usage, 'type': 'Device Usage'})


# ═══════════════════════════════════════════════════════════════════════════
# QR/BARCODE SCANNING SYSTEM
# ═══════════════════════════════════════════════════════════════════════════

import json
from django.apps import apps
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from .models import (
    ScanSession, ScanHistory, DeviceDailyUsageLog, DeviceUsageLogItem,
    DeviceTransferLog, PatientTransferLog, DeviceHandoverLog, DeviceAccessory, DeviceAccessoryUsageLog
)
from django.contrib.auth import get_user_model

User = get_user_model()

def parse_qr_code(qr_code, sample_token=None):
    """
    Parse QR code and return entity type, ID, and data
    Expected formats:
    - Standard: entity_type:id (e.g., "device:123", "bed:456")
    - Patient: patient:id|MRN:mrn|Name:first_last|DOB:yyyy-mm-dd
    - Operation tokens: op:operation_type (e.g., "op:usage", "op:transfer")
    - Lab tube: sample_token UUID (handled via sample_token parameter)
    """
    try:
        # Handle lab tube sample_token
        if sample_token:
            try:
                from laboratory.models import LabRequestItem
                lab_item = LabRequestItem.objects.select_related(
                    'request__patient', 'test'
                ).get(sample_token=sample_token)
                
                entity_data = {
                    'id': str(lab_item.sample_token),
                    'type': 'lab_tube',
                    'name': f"Lab Sample - {lab_item.test.english_name}",
                    'patient': str(lab_item.request.patient),
                    'patient_id': lab_item.request.patient.id,
                    'test_name': lab_item.test.english_name,
                    'status': lab_item.status,
                    'status_display': lab_item.get_status_display(),
                    'request_id': lab_item.request.id,
                    'sample_collected_at': lab_item.sample_collected_at.isoformat() if lab_item.sample_collected_at else None,
                    'sample_received_at': lab_item.sample_received_at.isoformat() if lab_item.sample_received_at else None,
                }
                
                return 'lab_tube', str(sample_token), entity_data, None
                
            except LabRequestItem.DoesNotExist:
                return None, None, None, f"Lab sample with token {sample_token} not found"
        
        if ':' not in qr_code:
            return None, None, None, "Invalid QR code format"
        
        # Check for operation tokens
        if qr_code.startswith('op:'):
            operation_type = qr_code.split(':', 1)[1]
            valid_operations = ['usage', 'transfer', 'handover', 'patient_transfer', 'clean', 'sterilize', 'inspect', 'maintenance']
            
            if operation_type in valid_operations:
                entity_data = {
                    'type': 'operation_token',
                    'operation': operation_type,
                    'name': f"Operation: {operation_type.title()}",
                    'description': f"Operation token for {operation_type}"
                }
                return 'operation_token', operation_type, entity_data, None
            else:
                return None, None, None, f"Unknown operation type: {operation_type}"
        
        # Check if it's a patient with extended format
        if qr_code.startswith('patient:') and '|' in qr_code:
            # Parse patient extended format: patient:id|MRN:mrn|Name:first_last|DOB:yyyy-mm-dd
            parts = qr_code.split('|')
            if len(parts) >= 4:
                try:
                    entity_type, entity_id = parts[0].split(':', 1)
                    entity_id = int(entity_id)
                    
                    # Extract additional patient data
                    mrn = parts[1].split(':', 1)[1] if len(parts[1].split(':', 1)) > 1 else ''
                    name = parts[2].split(':', 1)[1] if len(parts[2].split(':', 1)) > 1 else ''
                    dob = parts[3].split(':', 1)[1] if len(parts[3].split(':', 1)) > 1 else ''
                    
                    # Get patient from database
                    patient_model = apps.get_model('manager', 'Patient')
                    try:
                        entity = patient_model.objects.get(pk=entity_id)
                        
                        entity_data = {
                            'id': entity.pk,
                            'type': 'patient',
                            'name': str(entity),
                            'full_name': f"{entity.first_name} {entity.last_name}",
                            'mrn': entity.mrn,
                            'age': entity.age if hasattr(entity, 'age') else None,
                            'gender': getattr(entity, 'gender', None),
                            'date_of_birth': entity.date_of_birth.strftime('%Y-%m-%d') if entity.date_of_birth else dob,
                            'extra': {
                                'MRN': mrn,
                                'Name': name.replace('_', ' '),
                                'DOB': dob
                            }
                        }
                        
                        return 'patient', entity_id, entity_data, None
                        
                    except patient_model.DoesNotExist:
                        return None, None, None, f"Patient with ID {entity_id} not found"
                        
                except (ValueError, IndexError):
                    return None, None, None, "Invalid patient QR code format"
        
        # Standard format parsing
        entity_type, entity_id = qr_code.split(':', 1)
        entity_id = int(entity_id)
        
        # Map entity types to models
        model_mapping = {
            'device': ('maintenance', 'Device'),
            'patient': ('manager', 'Patient'),
            'bed': ('manager', 'Bed'),
            'user': ('hr', 'CustomUser'),
            'customuser': ('hr', 'CustomUser'),
            'accessory': ('maintenance', 'DeviceAccessory'),
            'deviceaccessory': ('maintenance', 'DeviceAccessory'),
            'department': ('manager', 'Department'),
            'doctor': ('manager', 'Doctor'),
        }
        
        if entity_type not in model_mapping:
            return None, None, None, f"Unknown entity type: {entity_type}"
        
        app_label, model_name = model_mapping[entity_type]
        model_class = apps.get_model(app_label, model_name)
        
        try:
            entity = model_class.objects.get(pk=entity_id)
            
            # Prepare entity data
            entity_data = {
                'id': entity.pk,
                'name': str(entity),
                'type': entity_type,
            }
            
            # Add specific fields based on entity type
            if entity_type == 'device':
                entity_data.update({
                    'status': entity.status,
                    'availability': entity.availability,
                    'clean_status': entity.clean_status,
                    'sterilization_status': entity.sterilization_status,
                    'department': str(entity.department) if entity.department else None,
                    'room': str(entity.room) if entity.room else None,
                    'bed': str(entity.bed) if entity.bed else None,
                    'current_patient': str(entity.current_patient) if entity.current_patient else None,
                })
            elif entity_type == 'patient':
                entity_data.update({
                    'full_name': f"{entity.first_name} {entity.last_name}",
                    'mrn': getattr(entity, 'mrn', ''),
                    'age': entity.age if hasattr(entity, 'age') else None,
                    'gender': getattr(entity, 'gender', None),
                    'date_of_birth': entity.date_of_birth.strftime('%Y-%m-%d') if entity.date_of_birth else None,
                })
            elif entity_type == 'bed':
                entity_data.update({
                    'status': entity.status,
                    'bed_type': entity.bed_type,
                    'bed_number': entity.bed_number,
                    'room': str(entity.room) if entity.room else None,
                    'department': str(entity.department) if hasattr(entity, 'department') else None,
                })
            elif entity_type in ['user', 'customuser']:
                entity_data.update({
                    'full_name': entity.get_full_name() if hasattr(entity, 'get_full_name') else str(entity),
                    'role': getattr(entity, 'role', None),
                    'email': getattr(entity, 'email', None),
                    'job_number': getattr(entity, 'job_number', None),
                })
            elif entity_type in ['accessory', 'deviceaccessory']:
                entity_data.update({
                    'device': str(entity.device) if entity.device else None,
                    'status': entity.status,
                    'serial_number': getattr(entity, 'serial_number', None),
                    'model': getattr(entity, 'model', None),
                })
            elif entity_type == 'doctor':
                entity_data.update({
                    'full_name': entity.user.get_full_name() if entity.user else str(entity),
                    'role': 'doctor',
                    'email': entity.user.email if entity.user else None,
                })
            
            return entity_type, entity_id, entity_data, None
            
        except model_class.DoesNotExist:
            return None, None, None, f"{entity_type.title()} with ID {entity_id} not found"
    
    except ValueError:
        return None, None, None, "Invalid entity ID format"
    except Exception as e:
        return None, None, None, f"Error parsing QR code: {str(e)}"


@login_required
def scan_qr_code(request):
    """
    API endpoint to scan and parse QR codes
    Enhanced for Step 3: Handles lab tubes and operation tokens
    """
    print(f"[DEBUG] scan_qr_code called - Method: {request.method}")
    print(f"[DEBUG] Request body: {request.body}")
    print(f"[DEBUG] User: {request.user}")
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body) if request.body else {}
        print(f"[DEBUG] Parsed data: {data}")
        qr_code = data.get('qr_code', '').strip()
        sample_token = data.get('sample_token', '').strip()
        session_id = data.get('session_id')
        print(f"[DEBUG] QR Code: {qr_code}, Session ID: {session_id}")
        
        if not qr_code and not sample_token:
            return JsonResponse({'error': 'QR code or sample token is required'}, status=400)
        
        # Parse QR code or sample token
        entity_type, entity_id, entity_data, error = parse_qr_code(qr_code, sample_token)
        
        if error:
            return JsonResponse({
                'success': False,
                'error': error,
                'qr_code': qr_code
            })
        
        # Get or create scan session
        scan_session = None
        if session_id:
            try:
                scan_session = ScanSession.objects.get(session_id=session_id, status='active')
            except ScanSession.DoesNotExist:
                pass
        
        # Create new session if none exists
        if not scan_session:
            scan_session = ScanSession.objects.create()
        
        # Add to scan history
        scan_history = ScanHistory.objects.create(
            session=scan_session,
            scanned_code=qr_code,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_data=entity_data,
            is_valid=True
        )
        
        # Handle session logic with operation auto-detection
        session_updated = False
        inferred_operation = None
        
        # Handle operation tokens
        if entity_type == 'operation_token':
            # Store operation hint in session context
            if not scan_session.context_json:
                scan_session.context_json = {}
            scan_session.context_json['operation_hint'] = entity_data['operation']
            session_updated = True
        
        # First scan should be a user (doctor/nurse/staff)
        elif not scan_session.user and entity_type in ['customuser', 'user', 'doctor']:
            if entity_type == 'doctor':
                # Get the user from doctor
                doctor_model = apps.get_model('manager', 'Doctor')
                doctor = doctor_model.objects.get(pk=entity_id)
                scan_session.user = doctor.user
            else:
                user_model = apps.get_model('hr', 'CustomUser')
                scan_session.user = user_model.objects.get(pk=entity_id)
            session_updated = True
        
        # Second scan can be patient or bed
        elif scan_session.user and not scan_session.patient and not scan_session.bed:
            if entity_type == 'patient':
                patient_model = apps.get_model('manager', 'Patient')
                scan_session.patient = patient_model.objects.get(pk=entity_id)
                session_updated = True
            elif entity_type == 'bed':
                bed_model = apps.get_model('manager', 'Bed')
                scan_session.bed = bed_model.objects.get(pk=entity_id)
                session_updated = True
        
        # Auto-detect operation type based on scan sequence
        scan_history = scan_session.scan_history.all()
        entity_types_scanned = [h.entity_type for h in scan_history]
        
        if 'user' in entity_types_scanned and 'patient' in entity_types_scanned and entity_type in ['device', 'accessory']:
            inferred_operation = 'usage'
        elif entity_type == 'device' and any(t in entity_types_scanned for t in ['department', 'room', 'bed']):
            inferred_operation = 'transfer'
        elif 'patient' in entity_types_scanned and entity_type == 'bed':
            inferred_operation = 'patient_transfer'
        elif entity_types_scanned.count('user') >= 2 and 'device' in entity_types_scanned:
            inferred_operation = 'handover'
        elif entity_type == 'lab_tube':
            inferred_operation = 'lab_sample_scan'
        
        if session_updated:
            scan_session.save()
        
        # Prepare response
        response_data = {
            'success': True,
            'session_id': str(scan_session.session_id),
            'entity_type': entity_type,
            'entity_id': entity_id,
            'entity_data': entity_data,
            'inferred_operation': inferred_operation,
            'session_status': {
                'user': scan_session.user.get_full_name() if scan_session.user else None,
                'patient': str(scan_session.patient) if scan_session.patient else None,
                'bed': str(scan_session.bed) if scan_session.bed else None,
                'scan_count': scan_session.scan_history.count(),
                'operation_hint': scan_session.context_json.get('operation_hint') if hasattr(scan_session, 'context_json') and scan_session.context_json else None,
            }
        }
        
        # Add validation warnings and notifications
        warnings = []
        notifications = []
        
        if entity_type == 'device':
            try:
                device = apps.get_model('maintenance', 'Device').objects.get(pk=entity_id)
                
                # Update device status when scanned for usage
                if inferred_operation == 'usage' and scan_session.user and scan_session.patient:
                    device.availability = False  # Mark as in use
                    device.save()
                    notifications.append(f"🔧 تم تعيين الجهاز كمستخدم: {device.name}")
                
                # Check if we have accessories scanned in this session that need linking
                accessory_scans = scan_session.scan_history.filter(entity_type='accessory')
                for accessory_scan in accessory_scans:
                    try:
                        accessory = apps.get_model('maintenance', 'DeviceAccessory').objects.get(pk=accessory_scan.entity_id)
                        
                        # Check if accessory is already linked to this device
                        if accessory.device_id == device.id:
                            notifications.append(f"✅ الملحق {accessory.name} مربوط بالفعل بالجهاز")
                        else:
                            # Check if accessory is linked to another device
                            if accessory.device and accessory.device != device:
                                # Create transfer request
                                transfer_request = apps.get_model('maintenance', 'AccessoryTransferRequest').objects.create(
                                    accessory=accessory,
                                    from_device=accessory.device,
                                    from_department=accessory.device.department,
                                    from_room=accessory.device.room,
                                    to_device=device,
                                    to_department=device.department,
                                    to_room=device.room,
                                    requested_by=scan_session.user,
                                    reason=f"نقل تلقائي من المسح - من {accessory.device.name} إلى {device.name}"
                                )
                                notifications.append(f"📋 تم إرسال طلب نقل الملحق {accessory.name} من {accessory.device.name} إلى {device.name}")
                                warnings.append(f"⚠️ الملحق {accessory.name} مربوط حالياً بجهاز آخر - تم إرسال طلب نقل")
                            else:
                                # Link accessory to device directly
                                old_device = accessory.device
                                accessory.device = device
                                accessory.save()
                                
                                # Create transfer log
                                apps.get_model('maintenance', 'AccessoryTransferLog').objects.create(
                                    accessory=accessory,
                                    from_device=old_device,
                                    from_department=old_device.department if old_device else None,
                                    from_room=old_device.room if old_device else None,
                                    to_device=device,
                                    to_department=device.department,
                                    to_room=device.room,
                                    transferred_by=scan_session.user,
                                    notes=f"ربط تلقائي من المسح"
                                )
                                notifications.append(f"✅ تم ربط الملحق {accessory.name} بالجهاز: {device.name}")
                    except apps.get_model('maintenance', 'DeviceAccessory').DoesNotExist:
                        continue
                
                if hasattr(device, 'status') and device.status == 'maintenance':
                    warnings.append(f"⚠️ الجهاز تحت الصيانة - لا يمكن استخدامه")
                elif hasattr(device, 'status') and device.status != 'working':
                    warnings.append(f"⚠️ الجهاز في حالة: {device.get_status_display()}")
                if hasattr(device, 'availability') and not device.availability:
                    warnings.append("⚠️ الجهاز غير متاح حالياً")
                if hasattr(device, 'clean_status') and device.clean_status == 'needs_cleaning':
                    warnings.append("⚠️ الجهاز يحتاج تنظيف")
                if hasattr(device, 'sterilization_status') and device.sterilization_status == 'needs_sterilization':
                    warnings.append("⚠️ الجهاز يحتاج تعقيم")
                notifications.append(f"✅ تم مسح الجهاز: {device.name}")
            except apps.get_model('maintenance', 'Device').DoesNotExist:
                warnings.append(f"⚠️ الجهاز رقم {entity_id} غير موجود في قاعدة البيانات")
                notifications.append(f"ℹ️ تم مسح كود جهاز: device:{entity_id}")
        
        elif entity_type == 'bed':
            try:
                bed = apps.get_model('manager', 'Bed').objects.get(pk=entity_id)
                if hasattr(bed, 'status') and bed.status == 'occupied':
                    warnings.append("⚠️ السرير مشغول حالياً")
                elif hasattr(bed, 'status') and bed.status == 'maintenance':
                    warnings.append("⚠️ السرير تحت الصيانة")
                notifications.append(f"✅ تم مسح السرير: {bed.name if hasattr(bed, 'name') else bed.id}")
            except apps.get_model('manager', 'Bed').DoesNotExist:
                warnings.append(f"⚠️ السرير رقم {entity_id} غير موجود في قاعدة البيانات")
                notifications.append(f"ℹ️ تم مسح كود سرير: bed:{entity_id}")
        
        elif entity_type == 'lab_tube':
            notifications.append(f"✅ تم مسح عينة المعمل: {entity_data['test_name']}")
            if entity_data['status'] == 'verified':
                notifications.append("✅ العينة معتمدة ومكتملة")
        
        elif entity_type == 'operation_token':
            notifications.append(f"🔧 تم تحديد نوع العملية: {entity_data['operation']}")
        
        elif entity_type == 'accessory':
            try:
                accessory = apps.get_model('maintenance', 'DeviceAccessory').objects.get(pk=entity_id)
                notifications.append(f"🔧 تم مسح الملحق: {accessory.name}")
                
                # Check if we have a device scanned in this session
                device_scans = scan_session.scan_history.filter(entity_type='device')
                if device_scans.exists():
                    # Get the last scanned device
                    last_device_scan = device_scans.last()
                    target_device = apps.get_model('maintenance', 'Device').objects.get(pk=last_device_scan.entity_id)
                    
                    # Check if accessory is already linked to this device
                    if accessory.device_id == target_device.id:
                        notifications.append(f"✅ الملحق مربوط بالفعل بالجهاز: {target_device.name}")
                    else:
                        # Check if accessory is linked to another device
                        if accessory.device and accessory.device != target_device:
                            # Create transfer request
                            transfer_request = apps.get_model('maintenance', 'AccessoryTransferRequest').objects.create(
                                accessory=accessory,
                                from_device=accessory.device,
                                from_department=accessory.device.department,
                                from_room=accessory.device.room,
                                to_device=target_device,
                                to_department=target_device.department,
                                to_room=target_device.room,
                                requested_by=scan_session.user,
                                reason=f"نقل تلقائي من المسح - من {accessory.device.name} إلى {target_device.name}"
                            )
                            notifications.append(f"📋 تم إرسال طلب نقل الملحق من {accessory.device.name} إلى {target_device.name}")
                            warnings.append(f"⚠️ الملحق مربوط حالياً بجهاز آخر - تم إرسال طلب نقل")
                        else:
                            # Link accessory to device directly
                            old_device = accessory.device
                            accessory.device = target_device
                            accessory.save()
                            
                            # Create transfer log
                            apps.get_model('maintenance', 'AccessoryTransferLog').objects.create(
                                accessory=accessory,
                                from_device=old_device,
                                from_department=old_device.department if old_device else None,
                                from_room=old_device.room if old_device else None,
                                to_device=target_device,
                                to_department=target_device.department,
                                to_room=target_device.room,
                                transferred_by=scan_session.user,
                                notes=f"ربط تلقائي من المسح"
                            )
                            notifications.append(f"✅ تم ربط الملحق بالجهاز: {target_device.name}")
                else:
                    notifications.append("ℹ️ امسح جهاز لربط الملحق به")
                    
            except apps.get_model('maintenance', 'DeviceAccessory').DoesNotExist:
                warnings.append(f"⚠️ الملحق رقم {entity_id} غير موجود في قاعدة البيانات")
                notifications.append(f"ℹ️ تم مسح كود ملحق: accessory:{entity_id}")
        
        elif entity_type == 'patient':
            notifications.append(f"👤 تم مسح المريض: {entity_data.get('name', entity_id)}")
        
        elif entity_type in ['user', 'customuser']:
            notifications.append(f"👨‍⚕️ تم مسح المستخدم: {entity_data.get('name', entity_id)}")
        
        elif entity_type == 'doctor':
            notifications.append(f"👨‍⚕️ تم مسح الطبيب: {entity_data.get('name', entity_id)}")
        
        else:
            notifications.append(f"✅ تم مسح الكود: {entity_type}:{entity_id}")
        
        response_data['warnings'] = warnings
        response_data['notifications'] = notifications
        
        return JsonResponse(response_data)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@login_required
def scan_session_page(request):
    """
    Main scanning session page - Step 3 Enhanced
    """
    # Get or create active session for user
    active_session = ScanSession.objects.filter(
        user=request.user, 
        status='active'
    ).first()
    
    if not active_session:
        active_session = ScanSession.objects.create(user=request.user)
    
    # Get recent scan history
    recent_scans = ScanHistory.objects.filter(
        session__user=request.user
    ).order_by('-scanned_at')[:20]
    
    # Get operation choices for dropdown
    operation_choices = [
        ('usage', 'استخدام الأجهزة'),
        ('transfer', 'نقل جهاز'),
        ('patient_transfer', 'نقل مريض'),
        ('handover', 'تسليم جهاز'),
        ('clean', 'تنظيف'),
        ('sterilize', 'تعقيم'),
        ('maintenance', 'صيانة'),
    ]
    
    context = {
        'session': active_session,
        'recent_scans': recent_scans,
        'operation_choices': operation_choices,
        'title': 'جلسة المسح المتكاملة - Step 3'
    }
    
    return render(request, 'maintenance/scan_session.html', context)


@csrf_exempt
@login_required
def save_scan_session(request):
    """
    Save the current scan session based on operation type
    Enhanced for Step 3: Handles all operation types
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        operation_type = data.get('operation_type', 'usage')
        notes = data.get('notes', '')
        
        if not session_id:
            return JsonResponse({'error': 'Session ID is required'}, status=400)
        
        # Get the scan session
        try:
            scan_session = ScanSession.objects.get(
                session_id=session_id, 
                status='active'
            )
        except ScanSession.DoesNotExist:
            return JsonResponse({'error': 'Session not found'}, status=404)
        
        created_records = []
        
        # Handle different operation types
        if operation_type == 'usage':
            # Create DeviceDailyUsageLog
            usage_log = DeviceDailyUsageLog.objects.create(
                user=scan_session.user or request.user,
                patient=scan_session.patient,
                bed=scan_session.bed,
                department=scan_session.bed.department if scan_session.bed and hasattr(scan_session.bed, 'department') else None,
                operation_type='surgery',  # Map 'usage' to valid choice
                notes=notes,
                is_completed=True,
                completed_at=timezone.now()
            )
            
            # Add scanned devices to the log
            device_scans = scan_session.scan_history.filter(entity_type='device')
            for scan in device_scans:
                try:
                    device = Device.objects.get(pk=scan.entity_id)
                    DeviceUsageLogItem.objects.create(
                        usage_log=usage_log,
                        device=device,
                        notes=f"Scanned at {scan.scanned_at.strftime('%H:%M:%S')}"
                    )
                    # Update device status
                    device.in_use = True
                    device.current_patient = scan_session.patient
                    device.usage_start_time = timezone.now()
                    device.save()
                except Device.DoesNotExist:
                    continue
            
            # Add scanned accessories
            accessory_scans = scan_session.scan_history.filter(entity_type='accessory')
            for scan in accessory_scans:
                try:
                    accessory = DeviceAccessory.objects.get(pk=scan.entity_id)
                    DeviceAccessoryUsageLog.objects.create(
                        usage_log=usage_log,
                        accessory=accessory,
                        notes=f"Scanned at {scan.scanned_at.strftime('%H:%M:%S')}"
                    )
                    # Update accessory status
                    accessory.status = 'in_use'
                    accessory.save()
                except DeviceAccessory.DoesNotExist:
                    continue
            
            created_records.append({'type': 'usage_log', 'id': usage_log.id})
        
        elif operation_type == 'transfer':
            # Handle device transfers
            device_scans = scan_session.scan_history.filter(entity_type='device')
            location_scans = scan_session.scan_history.filter(entity_type__in=['department', 'room', 'bed'])
            
            for device_scan in device_scans:
                try:
                    device = Device.objects.get(pk=device_scan.entity_id)
                    
                    # Get target location
                    to_department = None
                    to_room = None
                    to_bed = None
                    
                    for loc_scan in location_scans:
                        if loc_scan.entity_type == 'department':
                            to_department = Department.objects.get(pk=loc_scan.entity_id)
                        elif loc_scan.entity_type == 'room':
                            to_room = Room.objects.get(pk=loc_scan.entity_id)
                        elif loc_scan.entity_type == 'bed':
                            to_bed = Bed.objects.get(pk=loc_scan.entity_id)
                    
                    # Create transfer log
                    transfer_log = DeviceTransferLog.objects.create(
                        device=device,
                        from_department=device.department,
                        from_room=device.room,
                        from_bed=device.bed,
                        to_department=to_department or device.department,
                        to_room=to_room or device.room,
                        to_bed=to_bed,
                        moved_by=scan_session.user or request.user,
                        note=notes
                    )
                    
                    # Update device location
                    if to_department:
                        device.department = to_department
                    if to_room:
                        device.room = to_room
                    if to_bed:
                        device.bed = to_bed
                    device.save()
                    
                    created_records.append({'type': 'device_transfer', 'id': transfer_log.id})
                    
                except (Device.DoesNotExist, Department.DoesNotExist, Room.DoesNotExist, Bed.DoesNotExist):
                    continue
        
        elif operation_type == 'patient_transfer':
            # Handle patient transfers
            patient_scans = scan_session.scan_history.filter(entity_type='patient')
            bed_scans = scan_session.scan_history.filter(entity_type='bed')
            
            for patient_scan in patient_scans:
                for bed_scan in bed_scans:
                    try:
                        patient = Patient.objects.get(pk=patient_scan.entity_id)
                        to_bed = Bed.objects.get(pk=bed_scan.entity_id)
                        
                        # Get current patient location (from admission)
                        current_admission = patient.admission_set.filter(discharge_date__isnull=True).first()
                        
                        # Create transfer log
                        transfer_log = PatientTransferLog.objects.create(
                            patient=patient,
                            from_department=current_admission.department if current_admission else None,
                            from_bed=current_admission.bed if current_admission else None,
                            to_department=to_bed.department,
                            to_room=to_bed.room,
                            to_bed=to_bed,
                            moved_by=scan_session.user or request.user,
                            note=notes
                        )
                        
                        # Update patient location
                        if current_admission:
                            current_admission.bed = to_bed
                            current_admission.department = to_bed.department
                            current_admission.save()
                        
                        # Update bed status
                        to_bed.status = 'occupied'
                        to_bed.save()
                        
                        created_records.append({'type': 'patient_transfer', 'id': transfer_log.id})
                        
                    except (Patient.DoesNotExist, Bed.DoesNotExist):
                        continue
        
        elif operation_type == 'handover':
            # Handle device handovers
            device_scans = scan_session.scan_history.filter(entity_type='device')
            user_scans = scan_session.scan_history.filter(entity_type__in=['user', 'customuser', 'doctor'])
            
            if len(user_scans) >= 2:
                from_user_scan = user_scans[0]
                to_user_scan = user_scans[1]
                
                # Get users
                try:
                    if from_user_scan.entity_type == 'doctor':
                        from_user = Doctor.objects.get(pk=from_user_scan.entity_id).user
                    else:
                        from_user = User.objects.get(pk=from_user_scan.entity_id)
                    
                    if to_user_scan.entity_type == 'doctor':
                        to_user = Doctor.objects.get(pk=to_user_scan.entity_id).user
                    else:
                        to_user = User.objects.get(pk=to_user_scan.entity_id)
                    
                    for device_scan in device_scans:
                        try:
                            device = Device.objects.get(pk=device_scan.entity_id)
                            
                            # Create handover log
                            handover_log = DeviceHandoverLog.objects.create(
                                device=device,
                                from_user=from_user,
                                to_user=to_user,
                                note=notes
                            )
                            
                            created_records.append({'type': 'device_handover', 'id': handover_log.id})
                            
                        except Device.DoesNotExist:
                            continue
                            
                except (User.DoesNotExist, Doctor.DoesNotExist):
                    pass
        
        # Handle cleaning/sterilization/maintenance operations
        elif operation_type in ['clean', 'sterilize', 'maintenance']:
            device_scans = scan_session.scan_history.filter(entity_type='device')
            
            for device_scan in device_scans:
                try:
                    device = Device.objects.get(pk=device_scan.entity_id)
                    user = scan_session.user or request.user
                    
                    if operation_type == 'clean':
                        device.clean_status = 'clean'
                        device.last_cleaned_by = user
                        device.last_cleaned_at = timezone.now()
                        DeviceCleaningLog.objects.create(device=device, last_cleaned_by=user)
                    
                    elif operation_type == 'sterilize':
                        device.sterilization_status = 'sterilized'
                        device.last_sterilized_by = user
                        device.last_sterilized_at = timezone.now()
                        DeviceSterilizationLog.objects.create(device=device, last_sterilized_by=user)
                    
                    elif operation_type == 'maintenance':
                        device.status = 'working'
                        device.last_maintained_by = user
                        device.last_maintained_at = timezone.now()
                        DeviceMaintenanceLog.objects.create(device=device, last_maintained_by=user)
                    
                    device.save()
                    created_records.append({'type': f'device_{operation_type}', 'device_id': device.id})
                    
                except Device.DoesNotExist:
                    continue
        # Mark session as completed
        scan_session.status = 'completed'
        scan_session.save()
        
        return JsonResponse({
            'success': True,
            'operation_type': operation_type,
            'created_records': created_records,
            'message': f'تم حفظ عملية {operation_type} بنجاح'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@login_required
def device_usage_logs(request):
    """
    View device usage logs with filtering and export options
    """
    logs = DeviceUsageLogDaily.objects.all().select_related(
        'checked_by', 'device'
    )
    
    # Filtering
    user_filter = request.GET.get('user')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if user_filter:
        logs = logs.filter(checked_by_id=user_filter)
    if date_from:
        logs = logs.filter(date__gte=date_from)
    if date_to:
        logs = logs.filter(date__lte=date_to)
    
    # Get filter options
    users = User.objects.filter(deviceusagelogdaily__isnull=False).distinct()
    
    context = {
        'logs': logs.order_by('-date'),
        'users': users,
        'filters': {
            'user': user_filter,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    return render(request, 'maintenance/device_usage_logs.html', context)


@login_required
def export_device_usage_logs(request):
    """
    Export device usage logs to Excel/PDF
    """
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Get filtered logs (reuse filtering logic)
    logs = DeviceUsageLog.objects.all().select_related(
        'user', 'patient', 'bed', 'department'
    ).prefetch_related('items__device')
    
    # Apply same filters as in device_usage_logs view
    user_filter = request.GET.get('user')
    patient_filter = request.GET.get('patient')
    department_filter = request.GET.get('department')
    operation_filter = request.GET.get('operation_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if user_filter:
        logs = logs.filter(user_id=user_filter)
    if patient_filter:
        logs = logs.filter(patient_id=patient_filter)
    if department_filter:
        logs = logs.filter(department_id=department_filter)
    if operation_filter:
        logs = logs.filter(operation_type=operation_filter)
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Usage Logs"
    
    # Headers
    headers = [
        'Session ID', 'User', 'Patient', 'Bed', 'Department', 
        'Operation Type', 'Devices Used', 'Created At', 'Completed At', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, log in enumerate(logs.order_by('-created_at'), 2):
        devices_used = ', '.join([str(item.device) for item in log.items.all()])
        
        ws.cell(row=row, column=1, value=str(log.session_id))
        ws.cell(row=row, column=2, value=log.user.get_full_name())
        ws.cell(row=row, column=3, value=str(log.patient) if log.patient else '')
        ws.cell(row=row, column=4, value=str(log.bed) if log.bed else '')
        ws.cell(row=row, column=5, value=str(log.department) if log.department else '')
        ws.cell(row=row, column=6, value=log.get_operation_type_display())
        ws.cell(row=row, column=7, value=devices_used)
        ws.cell(row=row, column=8, value=log.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=9, value=log.completed_at.strftime('%Y-%m-%d %H:%M') if log.completed_at else '')
        ws.cell(row=row, column=10, value=log.notes or '')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="device_usage_logs_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3: ADDITIONAL API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════

@csrf_exempt
@login_required
def start_scan_session(request):
    """
    Start a new scan session
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body) if request.body else {}
        operation_type = data.get('operation_type')
        
        # Create new session
        session = ScanSession.objects.create(user=request.user)
        
        # Set operation hint if provided
        if operation_type:
            session.context_json = {'operation_hint': operation_type}
            session.save()
        
        return JsonResponse({
            'success': True,
            'session_id': str(session.session_id),
            'state': {
                'user': request.user.get_full_name(),
                'operation_hint': operation_type,
                'scan_count': 0
            }
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@csrf_exempt
@login_required
def reset_scan_session(request):
    """
    Reset/clear the current scan session
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body) if request.body else {}
        session_id = data.get('session_id')
        
        if session_id:
            try:
                session = ScanSession.objects.get(
                    session_id=session_id,
                    user=request.user,
                    status='active'
                )
                session.status = 'cancelled'
                session.save()
            except ScanSession.DoesNotExist:
                pass
        
        # Create new session
        new_session = ScanSession.objects.create(user=request.user)
        
        return JsonResponse({
            'success': True,
            'session_id': str(new_session.session_id),
            'message': 'تم إعادة تعيين الجلسة'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@login_required
def get_session_status(request, session_id):
    """
    Get current session status and scan history
    """
    try:
        session = ScanSession.objects.get(
            session_id=session_id,
            user=request.user
        )
        
        scan_history = session.scan_history.all().order_by('-scanned_at')
        
        return JsonResponse({
            'success': True,
            'session': {
                'id': str(session.session_id),
                'status': session.status,
                'user': session.user.get_full_name() if session.user else None,
                'patient': str(session.patient) if session.patient else None,
                'bed': str(session.bed) if session.bed else None,
                'created_at': session.created_at.isoformat(),
                'scan_count': scan_history.count(),
            },
            'scan_history': [{
                'entity_type': scan.entity_type,
                'entity_data': scan.entity_data,
                'scanned_at': scan.scanned_at.isoformat(),
                'is_valid': scan.is_valid
            } for scan in scan_history[:20]]  # Last 20 scans
        })
    
    except ScanSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


def qr_test_page(request):
    """
    QR Test page for Step 2 implementation
    Test scanning of Patient, Bed, CustomUser, and DeviceAccessory QR codes
    """
    return render(request, 'maintenance/qr_test.html', {
        'title': 'اختبار نظام QR/Barcode - Step 2'
    })


@csrf_exempt
@login_required
def generate_qr_code(request):
    """
    Generate QR code for any entity (bed, device, patient, etc.)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        entity_type = data.get('entity_type')
        entity_id = data.get('entity_id')
        
        if not entity_type or not entity_id:
            return JsonResponse({'error': 'entity_type and entity_id are required'}, status=400)
        
        # Map entity types to models
        model_mapping = {
            'bed': apps.get_model('manager', 'Bed'),
            'device': apps.get_model('maintenance', 'Device'),
            'accessory': apps.get_model('maintenance', 'DeviceAccessory'),
            'patient': apps.get_model('manager', 'Patient'),
            'user': apps.get_model('hr', 'CustomUser'),
        }
        
        if entity_type not in model_mapping:
            return JsonResponse({'error': f'Unsupported entity type: {entity_type}'}, status=400)
        
        model_class = model_mapping[entity_type]
        
        try:
            entity = model_class.objects.get(pk=entity_id)
        except model_class.DoesNotExist:
            return JsonResponse({'error': f'{entity_type} with ID {entity_id} not found'}, status=404)
        
        # Generate QR code
        if hasattr(entity, 'generate_qr_code'):
            entity.generate_qr_code()
            entity.save()
            
            return JsonResponse({
                'success': True,
                'qr_code_url': entity.qr_code.url if entity.qr_code else None,
                'qr_token': entity.qr_token
            })
        else:
            return JsonResponse({'error': f'{entity_type} does not support QR code generation'}, status=400)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


def test_links(request):
    return render(request, "maintenance/test_links.html")